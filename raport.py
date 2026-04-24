import yfinance as yf
import pandas as pd
from datetime import date, timedelta

# Portfel spółek GPW
SPOLKI = {
    "PKO BP":     "PKO.WA",
    "CD Projekt":  "CDR.WA",
    "ORLEN":      "PKN.WA",
    "Allegro":    "ALE.WA",
    "Dino":       "DNP.WA"
}

# Okres analizy — ostatnie 3 miesiące
koniec = date.today()
poczatek = koniec - timedelta(days=90)

print("Pobieranie danych...")

dane = {}
for nazwa, ticker in SPOLKI.items():
    df = yf.Ticker(ticker).history(
        start=poczatek.strftime("%Y-%m-%d"),
        end=koniec.strftime("%Y-%m-%d")
    )
    if not df.empty:
        dane[nazwa] = df["Close"]
        print(f"  ✅ {nazwa} ({ticker}) — {len(df)} sesji")
    else:
        print(f"  ❌ {nazwa} ({ticker}) — brak danych")

# Połącz w jedną tabelę
df_ceny = pd.DataFrame(dane)
df_ceny.index = df_ceny.index.tz_convert(None)
df_ceny.index = df_ceny.index.date

print(f"\nPobrano dane: {len(df_ceny)} sesji, {len(df_ceny.columns)} spółek")
print(df_ceny.tail(3))

# ── MODUŁ 2 — Statystyki ──────────────────────────────────────

# Zmiana procentowa: pierwszy dzień → ostatni dzień okresu
cena_poczatkowa = df_ceny.iloc[0]
cena_koncowa    = df_ceny.iloc[-1]
zmiana_proc     = ((cena_koncowa - cena_poczatkowa) / cena_poczatkowa * 100).round(2)

# Statystyki opisowe
statystyki = pd.DataFrame({
    "Cena początkowa": cena_poczatkowa.round(2),
    "Cena końcowa":    cena_koncowa.round(2),
    "Zmiana %":        zmiana_proc,
    "Min (3M)":        df_ceny.min().round(2),
    "Max (3M)":        df_ceny.max().round(2),
    "Średnia (3M)":    df_ceny.mean().round(2),
})

print("\n── Statystyki portfela (ostatnie 3 miesiące) ──")
print(statystyki.to_string())

# ── MODUŁ 3 — Ocena portfela ─────────────────────────────────

print("\n── Ocena portfela ──")

for spolka, zmiana in zmiana_proc.items():
    if zmiana >= 10:
        ocena = "🚀 Silny wzrost"
    elif zmiana >= 0:
        ocena = "📈 Wzrost"
    elif zmiana >= -10:
        ocena = "📉 Spadek"
    else:
        ocena = "🔴 Silny spadek"

    print(f"  {spolka:<12} {zmiana:>7.2f}%   {ocena}")

# Podsumowanie
zyskaly  = (zmiana_proc > 0).sum()
stracily = (zmiana_proc < 0).sum()
srednia_portfela = zmiana_proc.mean().round(2)

print(f"\n  Spółek na plusie:  {zyskaly}")
print(f"  Spółek na minusie: {stracily}")
print(f"  Średnia zmiana portfela: {srednia_portfela}%")

# ── MODUŁ 4 — Eksport do Excela ──────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

nazwa_pliku = "raport_gpw.xlsx"
wb = Workbook()
ws = wb.active
ws.title = "Analiza portfela"

# Kolory
ZIELONY    = PatternFill("solid", fgColor="C6EFCE")
CZERWONY   = PatternFill("solid", fgColor="FFC7CE")
ZOLTY      = PatternFill("solid", fgColor="FFEB9C")
NAGLOWEK   = PatternFill("solid", fgColor="1F4E79")
CZCIONKA_B = Font(bold=True, color="FFFFFF")
CZCIONKA_N = Font(bold=True)

# Nagłówek raportu
ws.merge_cells("A1:G1")
ws["A1"] = f"Raport GPW — analiza portfela | {koniec.strftime('%d.%m.%Y')}"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill = NAGLOWEK
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Nagłówki kolumn
naglowki = ["Spółka", "Cena pocz.", "Cena końc.", "Zmiana %", "Min (3M)", "Max (3M)", "Średnia (3M)"]
for col, naglowek in enumerate(naglowki, 1):
    komorka = ws.cell(row=2, column=col, value=naglowek)
    komorka.fill = NAGLOWEK
    komorka.font = CZCIONKA_B
    komorka.alignment = Alignment(horizontal="center")

# Dane spółek
for row, (spolka, dane_row) in enumerate(statystyki.iterrows(), 3):
    ws.cell(row=row, column=1, value=spolka).font = CZCIONKA_N
    ws.cell(row=row, column=2, value=float(dane_row["Cena początkowa"]))
    ws.cell(row=row, column=3, value=float(dane_row["Cena końcowa"]))
    
    zmiana_komorka = ws.cell(row=row, column=4, value=float(dane_row["Zmiana %"]))
    zmiana_komorka.number_format = '+0.00"%";-0.00"%"'      
    # Kolorowanie zmiany %
    val = float(dane_row["Zmiana %"])
    if val >= 10:
        zmiana_komorka.fill = ZIELONY
    elif val >= 0:
        zmiana_komorka.fill = PatternFill("solid", fgColor="E2EFDA")
    elif val >= -10:
        zmiana_komorka.fill = ZOLTY
    else:
        zmiana_komorka.fill = CZERWONY

    ws.cell(row=row, column=5, value=float(dane_row["Min (3M)"]))
    ws.cell(row=row, column=6, value=float(dane_row["Max (3M)"]))
    ws.cell(row=row, column=7, value=float(dane_row["Średnia (3M)"]))

# Podsumowanie
wiersz_sum = len(statystyki) + 4
ws.cell(row=wiersz_sum,   column=1, value="Spółek na plusie:").font  = CZCIONKA_N
ws.cell(row=wiersz_sum,   column=2, value=int(zyskaly))
ws.cell(row=wiersz_sum+1, column=1, value="Spółek na minusie:").font = CZCIONKA_N
ws.cell(row=wiersz_sum+1, column=2, value=int(stracily))
ws.cell(row=wiersz_sum+2, column=1, value="Średnia zmiana portfela:").font = CZCIONKA_N
ws.cell(row=wiersz_sum+2, column=2, value=float(srednia_portfela))

# Szerokości kolumn
szerokosci = [16, 12, 12, 12, 12, 12, 14]
for i, sz in enumerate(szerokosci, 1):
    ws.column_dimensions[get_column_letter(i)].width = sz

# ── MODUŁ 5 — Wykres ─────────────────────────────────────────
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.axis import ChartLines

wykres = BarChart()
wykres.type = "col"
wykres.grouping = "clustered"
wykres.title = "Zmiana % spółek GPW — ostatnie 3 miesiące"
wykres.style = 10
wykres.width = 22
wykres.height = 14

# Oś Y
wykres.y_axis.title = "Zmiana %"
wykres.y_axis.scaling.min = -20
wykres.y_axis.scaling.max = 30
wykres.y_axis.majorUnit = 5
wykres.y_axis.numFmt = '0"%"'
wykres.y_axis.majorGridlines = ChartLines()

# Oś X
wykres.x_axis.title = "Spółka"
wykres.x_axis.tickLblPos = "low"    # etykiety spółek na dole
wykres.x_axis.crosses = "autoZero"  # linia na poziomie zera

# Seria danych
dane_wykresu = Reference(ws, min_col=4, min_row=3, max_row=len(statystyki)+2)
kategorie    = Reference(ws, min_col=1, min_row=3, max_row=len(statystyki)+2)

seria = Series(dane_wykresu, title="Zmiana %")
seria.invertIfNegative = False

# Etykiety na słupkach
seria.dLbls = DataLabelList()
seria.dLbls.showVal = True
seria.dLbls.showLegendKey = False
seria.dLbls.showCatName = False
seria.dLbls.showSerName = False
seria.dLbls.numFmt = '0.00"%"'

wykres.append(seria)
wykres.set_categories(kategorie)

ws.add_chart(wykres, "A13")

wb.save(nazwa_pliku)

# — MODUŁ X — Generowanie strony HTML —————————————————

def generuj_html(statystyki, poczatek, koniec):
    
    wiersze = ""
    for spolka, row in statystyki.iterrows():
        zmiana = row["Zmiana %"]
        kolor = "#3fb950" if zmiana >= 0 else "#f85149"  # zielony / czerwony
        znak = "+" if zmiana >= 0 else ""
        wiersze += f"""
        <tr>
            <td>{spolka}</td>
            <td>{row['Cena początkowa']:.2f} zł</td>
            <td>{row['Cena końcowa']:.2f} zł</td>
            <td style="color:{kolor}; font-weight:bold">{znak}{zmiana:.2f}%</td>
            <td>{row['Min (3M)']:.2f} zł</td>
            <td>{row['Max (3M)']:.2f} zł</td>
            <td>{row['Średnia (3M)']:.2f} zł</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="pl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Raport GPW — aitomek</title>
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            max-width: 1000px;
            margin: 40px auto;
            padding: 0 20px;
            background: #0d1117;
            color: #e6edf3;
        }}
        h1 {{ color: #58a6ff; border-bottom: 1px solid #30363d; padding-bottom: 12px; }}
        h2 {{ color: #58a6ff; margin-top: 32px; }}
        .meta {{ color: #8b949e; font-size: 14px; margin-bottom: 32px; }}
        .badge {{
            display: inline-block;
            padding: 3px 10px;
            border-radius: 12px;
            font-size: 12px;
            background: #1f6feb;
            color: white;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 16px;
        }}
        th {{
            background: #161b22;
            padding: 10px 14px;
            text-align: left;
            border: 1px solid #30363d;
            color: #58a6ff;
        }}
        td {{
            padding: 10px 14px;
            border: 1px solid #30363d;
        }}
        tr:hover {{ background: #161b22; }}
        .footer {{
            margin-top: 48px;
            color: #8b949e;
            font-size: 13px;
            border-top: 1px solid #30363d;
            padding-top: 16px;
        }}
    </style>
</head>
<body>

<h1>📊 Raport Analizy Portfela GPW</h1>
<p class="meta">
    Dane: {poczatek.strftime('%d.%m.%Y')} – {koniec.strftime('%d.%m.%Y')} · 
    Aktualizacja: {koniec.strftime('%d.%m.%Y')} · 
    <span class="badge">automatyczny</span>
</p>

<h2>Statystyki portfela</h2>
<table>
    <tr>
        <th>Spółka</th>
        <th>Cena pocz.</th>
        <th>Cena końc.</th>
        <th>Zmiana %</th>
        <th>Min (3M)</th>
        <th>Max (3M)</th>
        <th>Średnia (3M)</th>
    </tr>
    {wiersze}
</table>

<div class="footer">
    Źródło danych: Yahoo Finance · 
    Kod: <a href="https://github.com/aitomek01/raport-gpw" style="color:#58a6ff">
    github.com/aitomek01/raport-gpw</a>
</div>

</body>
</html>"""

    with open("index.html", "w", encoding="utf-8") as f:
        f.write(html)
    print("✅ Wygenerowano index.html")

generuj_html(statystyki, poczatek, koniec)

# ── Usuń siatkę z wykresu (patch XML) ────────────────────────
import zipfile, os

def usun_siatke(plik):
    tmp = plik + ".tmp"
    with zipfile.ZipFile(plik, 'r') as zin:
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith('xl/charts/'):
                    # Obsługa obu wariantów — ze spacją i bez
                    data = data.replace(b'<c:majorGridlines />', b'')
                    data = data.replace(b'<c:majorGridlines/>', b'')
                    data = data.replace(b'<majorGridlines />', b'')
                    data = data.replace(b'<majorGridlines/>', b'')
                zout.writestr(item, data)
    os.replace(tmp, plik)

usun_siatke(nazwa_pliku)
print("✅ Siatka usunięta")